#!/usr/bin/env python3
"""Shared, bounded readers for the CXBA Sandbox material scripts."""

from __future__ import annotations

import csv
import json
import os
import tempfile
from collections.abc import Iterator, Mapping, Sequence
from pathlib import Path
from typing import Any, TextIO


ALLOWED_ACCOUNT_TYPES = frozenset({"BANK_ACCOUNT", "BANK_CARD", "WECHAT", "ALIPAY"})


class MaterialToolError(RuntimeError):
    """A safe, user-actionable material tool failure."""


def atomic_write_text(path: Path, content: str) -> None:
    path = path.resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp")
    temporary.write_text(content, encoding="utf-8")
    temporary.replace(path)


def atomic_write_lines(path: Path, lines: Iterator[str]) -> None:
    """Write an arbitrarily large text result without building it in memory."""

    path = path.resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8") as handle:
            for line in lines:
                handle.write(line)
        temporary.replace(path)
    except BaseException:
        temporary.unlink(missing_ok=True)
        raise


def json_default(value: Any) -> Any:
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if hasattr(value, "item"):
        return value.item()
    raise TypeError(f"Unsupported JSON value type: {type(value).__name__}")


def write_json(path: Path, payload: Any) -> None:
    atomic_write_text(
        path,
        json.dumps(payload, ensure_ascii=False, indent=2, default=json_default) + "\n",
    )


def read_json(path: Path) -> Any:
    try:
        with path.open("r", encoding="utf-8") as handle:
            return json.load(handle)
    except (OSError, json.JSONDecodeError) as exc:
        raise MaterialToolError(f"Cannot read JSON input: {type(exc).__name__}: {exc}") from exc


def safe_relative_path(raw_path: str) -> Path:
    candidate = Path(str(raw_path).replace("\\", "/"))
    if candidate.is_absolute() or not candidate.parts or ".." in candidate.parts:
        raise MaterialToolError("Material relativePath must stay below the data root")
    return candidate


def resolve_below(root: Path, relative_path: str) -> Path:
    root = root.resolve(strict=True)
    candidate = (root / safe_relative_path(relative_path)).resolve(strict=True)
    if candidate == root or root not in candidate.parents or not candidate.is_file():
        raise MaterialToolError("Material path does not resolve to a file below the data root")
    return candidate


def catalog_entries(payload: Any) -> list[Mapping[str, Any]]:
    if isinstance(payload, list):
        entries = payload
    elif isinstance(payload, Mapping):
        entries = payload.get("materials") or payload.get("files") or payload.get("items")
    else:
        entries = None
    if not isinstance(entries, list):
        raise MaterialToolError("Catalog must be a list or contain materials/files/items")
    if not all(isinstance(entry, Mapping) for entry in entries):
        raise MaterialToolError("Every catalog entry must be an object")
    return entries


def entry_value(entry: Mapping[str, Any], *names: str) -> Any:
    for name in names:
        if name in entry:
            return entry[name]
    return None


def load_material(catalog_path: Path, material_id: str) -> tuple[Mapping[str, Any], Path]:
    payload = read_json(catalog_path)
    if not isinstance(payload, Mapping):
        raise MaterialToolError("Normalized catalog must be an object")
    root_value = payload.get("dataRoot")
    if not isinstance(root_value, str) or not root_value:
        raise MaterialToolError("Normalized catalog is missing dataRoot")
    matches = [
        entry
        for entry in catalog_entries(payload)
        if str(entry_value(entry, "materialId", "id") or "") == material_id
    ]
    if len(matches) != 1:
        raise MaterialToolError(f"Expected one catalog entry for materialId, found {len(matches)}")
    entry = matches[0]
    relative_path = entry_value(entry, "relativePath", "path")
    if not isinstance(relative_path, str) or not relative_path:
        raise MaterialToolError("Material entry is missing relativePath")
    return entry, resolve_below(Path(root_value), relative_path)


def sanitized_name(value: str) -> str:
    safe = "".join(char if char.isalnum() or char in "-_." else "_" for char in value)
    return safe.strip("._") or "material"


def _open_csv(path: Path) -> tuple[TextIO, csv.Dialect]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "gb18030"):
        handle: TextIO | None = None
        try:
            handle = path.open("r", encoding=encoding, newline="")
            sample = handle.read(65536)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
            except csv.Error:
                dialect = csv.excel_tab if path.suffix.lower() == ".tsv" else csv.excel
            return handle, dialect
        except UnicodeError as exc:
            last_error = exc
            if handle is not None:
                handle.close()
    raise MaterialToolError(f"Cannot decode delimited text: {last_error}")


def list_tables(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        return ["data"]
    if suffix in {".xlsx", ".xlsm"}:
        import openpyxl

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()
    if suffix == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(path, on_demand=True)
        try:
            return list(workbook.sheet_names())
        finally:
            workbook.release_resources()
    if suffix in {".parquet", ".pq"}:
        return ["data"]
    if suffix in {".duckdb", ".ddb"}:
        import duckdb

        connection = duckdb.connect(str(path), read_only=True)
        try:
            return [row[0] for row in connection.execute("SHOW TABLES").fetchall()]
        finally:
            connection.close()
    raise MaterialToolError(f"Unsupported tabular format: {suffix or '<none>'}")


def _unique_headers(values: Sequence[Any]) -> list[str]:
    seen: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(values, start=1):
        base = str(value).strip() if value is not None and str(value).strip() else f"column_{index}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def _row_mapping(headers: Sequence[str], values: Sequence[Any]) -> dict[str, Any]:
    padded = list(values[: len(headers)]) + [None] * max(0, len(headers) - len(values))
    return dict(zip(headers, padded, strict=True))


def iter_table_rows(
    path: Path,
    *,
    table: str | None = None,
    header_row: int = 1,
    batch_size: int = 2048,
) -> tuple[list[str], Iterator[tuple[int, dict[str, Any]]]]:
    """Return headers and a streaming iterator of one-based source rows."""

    if header_row < 1:
        raise MaterialToolError("header_row must be at least 1")
    suffix = path.suffix.lower()

    if suffix in {".csv", ".tsv", ".txt"}:
        handle, dialect = _open_csv(path)
        reader = csv.reader(handle, dialect)
        try:
            for _ in range(header_row - 1):
                next(reader)
            headers = _unique_headers(next(reader))
        except StopIteration as exc:
            handle.close()
            raise MaterialToolError("Header row is outside the delimited file") from exc

        def csv_rows() -> Iterator[tuple[int, dict[str, Any]]]:
            try:
                for row_number, values in enumerate(reader, start=header_row + 1):
                    yield row_number, _row_mapping(headers, values)
            finally:
                handle.close()

        return headers, csv_rows()

    if suffix in {".xlsx", ".xlsm"}:
        import openpyxl

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        selected = table or workbook.sheetnames[0]
        if selected not in workbook.sheetnames:
            workbook.close()
            raise MaterialToolError(f"Worksheet not found: {selected}")
        rows = workbook[selected].iter_rows(values_only=True)
        try:
            for _ in range(header_row - 1):
                next(rows)
            headers = _unique_headers(next(rows))
        except StopIteration as exc:
            workbook.close()
            raise MaterialToolError("Header row is outside the worksheet") from exc

        def xlsx_rows() -> Iterator[tuple[int, dict[str, Any]]]:
            try:
                for row_number, values in enumerate(rows, start=header_row + 1):
                    yield row_number, _row_mapping(headers, values)
            finally:
                workbook.close()

        return headers, xlsx_rows()

    if suffix == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(path, on_demand=True)
        selected = table or workbook.sheet_names()[0]
        try:
            sheet = workbook.sheet_by_name(selected)
        except xlrd.biffh.XLRDError as exc:
            workbook.release_resources()
            raise MaterialToolError(f"Worksheet not found: {selected}") from exc
        if header_row > sheet.nrows:
            workbook.release_resources()
            raise MaterialToolError("Header row is outside the worksheet")
        headers = _unique_headers(sheet.row_values(header_row - 1))

        def xls_rows() -> Iterator[tuple[int, dict[str, Any]]]:
            try:
                for zero_index in range(header_row, sheet.nrows):
                    yield zero_index + 1, _row_mapping(headers, sheet.row_values(zero_index))
            finally:
                workbook.release_resources()

        return headers, xls_rows()

    if suffix in {".parquet", ".pq"}:
        import pyarrow.parquet as parquet

        parquet_file = parquet.ParquetFile(path)
        headers = list(parquet_file.schema_arrow.names)

        def parquet_rows() -> Iterator[tuple[int, dict[str, Any]]]:
            row_number = 0
            for batch in parquet_file.iter_batches(batch_size=batch_size):
                for row in batch.to_pylist():
                    row_number += 1
                    yield row_number, row

        return headers, parquet_rows()

    if suffix in {".duckdb", ".ddb"}:
        import duckdb

        if not table:
            raise MaterialToolError("DuckDB input requires --sheet with a table name")
        connection = duckdb.connect(str(path), read_only=True)
        escaped = table.replace('"', '""')
        cursor = connection.execute(f'SELECT * FROM "{escaped}"')
        headers = [description[0] for description in cursor.description]

        def duckdb_rows() -> Iterator[tuple[int, dict[str, Any]]]:
            row_number = 0
            try:
                while True:
                    batch = cursor.fetchmany(batch_size)
                    if not batch:
                        break
                    for values in batch:
                        row_number += 1
                        yield row_number, _row_mapping(headers, values)
            finally:
                connection.close()

        return headers, duckdb_rows()

    raise MaterialToolError(f"Unsupported tabular format: {suffix or '<none>'}")


def require_columns(headers: Sequence[str], columns: Sequence[str]) -> None:
    missing = sorted(set(columns) - set(headers))
    if missing:
        raise MaterialToolError(f"Columns not found: {', '.join(missing)}")

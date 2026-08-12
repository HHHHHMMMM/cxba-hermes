#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

SUPPORTED = {".xlsx", ".xls", ".csv", ".tsv"}
MAX_HEADER_SCAN_ROWS = 20
LONG_NUMBER = re.compile(r"(?<!\d)\d{7,}(?!\d)")
EMAIL = re.compile(r"[^\s@]+@[^\s@]+\.[^\s@]+")
HEADER_HINTS = (
    "姓名", "名称", "日期", "时间", "金额", "账户", "账号", "卡号", "证件", "摘要",
    "对方", "对手", "来源", "收方", "付方", "申请", "部门", "单据", "交易", "流水", "客户",
    "余额", "备注", "序号", "编号", "机构", "科目", "费用", "类型", "状态",
    "户名", "主体", "人员", "企业", "单位", "银行", "币种", "借贷", "方向",
    "用途", "渠道", "柜员", "网点", "地址", "电话", "邮箱", "职业", "职务",
    "关系", "比例", "次数", "笔数", "数量", "期间", "年度", "月份", "代码",
    "标志", "凭证", "发票", "税号", "成本", "预算", "结算", "审批", "收款",
    "付款", "收入", "支出", "发生额", "借方", "贷方", "IP", "MAC",
)


def fail(message: str) -> None:
    raise SystemExit(message)


def safe_paths(root_value: str, output_value: str) -> tuple[Path, Path]:
    root = Path(root_value).expanduser().resolve(strict=True)
    output = Path(output_value).expanduser().resolve()
    if not root.is_dir():
        fail("ROOT_NOT_DIRECTORY")
    if output == root or root in output.parents:
        fail("OUTPUT_MUST_BE_OUTSIDE_SOURCE_ROOT")
    output.parent.mkdir(parents=True, exist_ok=True)
    return root, output


def clean_header(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float, date, datetime)):
        return "[NON_HEADER_VALUE]"
    text = str(value).strip()
    if not text:
        return ""
    if len(text) > 80 or LONG_NUMBER.search(text) or EMAIL.search(text):
        return "[NON_HEADER_VALUE]"
    return text


def choose_header(rows: Iterable[Iterable[Any]]) -> tuple[int | None, list[str]]:
    best_index = None
    best_values: list[str] = []
    best_score = -1
    for index, row in enumerate(rows, start=1):
        values = [clean_header(value) for value in row]
        nonempty = [value for value in values if value]
        hint_score = sum(
            any(hint in value for hint in HEADER_HINTS)
            for value in nonempty
            if not value.startswith("[")
        )
        score = hint_score * 100 + min(len(nonempty), 20) - sum(
            value.startswith("[") for value in nonempty
        ) * 10
        if score > best_score:
            best_index = index
            best_values = values
            best_score = score
    while best_values and not best_values[-1]:
        best_values.pop()
    safe_values = []
    for column_index, value in enumerate(best_values, start=1):
        normalized = re.sub(r"\s+", "", value)
        english_header = bool(
            re.fullmatch(r"[A-Z][A-Z0-9_.-]{1,50}", normalized)
            or ("_" in normalized and re.fullmatch(r"[A-Za-z][A-Za-z0-9_.-]{1,50}", normalized))
        )
        if any(hint in normalized for hint in HEADER_HINTS) or english_header:
            safe_values.append(value)
        elif value:
            safe_values.append(f"COLUMN_{column_index}")
        else:
            safe_values.append("")
    return best_index, safe_values


def safe_relative_path(relative: str) -> str:
    return relative


def xlsx_sheets(path: Path) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError:
        fail("DEPENDENCY_MISSING: openpyxl")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = []
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(
                min_row=1,
                max_row=min(sheet.max_row, MAX_HEADER_SCAN_ROWS),
                values_only=True,
            )
            header_row, headers = choose_header(rows)
            result.append({
                "name": sheet.title,
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "headerRow": header_row,
                "headers": headers,
            })
    finally:
        workbook.close()
    return result


def xls_sheets(path: Path) -> list[dict[str, Any]]:
    try:
        import xlrd
    except ImportError:
        fail("DEPENDENCY_MISSING: xlrd")
    workbook = xlrd.open_workbook(path, on_demand=True)
    result = []
    try:
        for sheet in workbook.sheets():
            rows = (
                sheet.row_values(index)
                for index in range(min(sheet.nrows, MAX_HEADER_SCAN_ROWS))
            )
            header_row, headers = choose_header(rows)
            result.append({
                "name": sheet.name,
                "rows": sheet.nrows,
                "columns": sheet.ncols,
                "headerRow": header_row,
                "headers": headers,
            })
            workbook.unload_sheet(sheet.name)
    finally:
        workbook.release_resources()
    return result


def delimited_schema(path: Path, delimiter: str) -> list[dict[str, Any]]:
    rows: list[list[str]] = []
    row_count = 0
    encoding_used = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="", errors="strict") as handle:
                reader = csv.reader(handle, delimiter=delimiter)
                for row_count, row in enumerate(reader, start=1):
                    if row_count <= MAX_HEADER_SCAN_ROWS:
                        rows.append(row)
            encoding_used = encoding
            break
        except UnicodeDecodeError:
            rows = []
            row_count = 0
    if encoding_used is None:
        return [{"name": "", "error": "UNSUPPORTED_ENCODING"}]
    header_row, headers = choose_header(rows)
    return [{
        "name": "",
        "rows": row_count,
        "columns": max((len(row) for row in rows), default=0),
        "headerRow": header_row,
        "headers": headers,
        "encoding": encoding_used,
    }]


def describe(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return xlsx_sheets(path)
    if suffix == ".xls":
        return xls_sheets(path)
    if suffix == ".csv":
        return delimited_schema(path, ",")
    if suffix == ".tsv":
        return delimited_schema(path, "\t")
    return []


def files_under(root: Path) -> list[Path]:
    return sorted(path for path in root.rglob("*") if path.is_file())


def inventory(root: Path) -> dict[str, Any]:
    records = []
    for file_index, path in enumerate(files_under(root), start=1):
        file_id = f"F{file_index:04d}"
        relative = path.relative_to(root).as_posix()
        record: dict[str, Any] = {
            "fileId": file_id,
            "path": safe_relative_path(relative),
            "extension": path.suffix.lower(),
            "sizeBytes": path.stat().st_size,
        }
        if path.suffix.lower() in SUPPORTED:
            try:
                record["sheets"] = describe(path)
            except Exception as error:
                record["parseError"] = type(error).__name__
        records.append(record)
    return {
        "rootLabel": root.name,
        "fileCount": len(records),
        "supportedFileCount": sum(
            record["extension"] in SUPPORTED for record in records
        ),
        "files": records,
    }


def header_name(headers: list[str], column_index: int) -> str:
    if 0 <= column_index < len(headers) and headers[column_index]:
        return headers[column_index]
    return f"COLUMN_{column_index + 1}"


def json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, float) and not math.isfinite(value):
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def value_type(value: Any) -> str:
    if value is None or value == "":
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "number"
    return "text"


def content_profile(
    row_stream: Iterable[tuple[int, Iterable[Any]]],
    header_row: int | None,
    headers: list[str],
    total_rows: int,
    total_columns: int,
) -> dict[str, Any]:
    counts = [0] * total_columns
    types: list[dict[str, int]] = [{} for _ in range(total_columns)]
    samples = []
    data_rows = 0
    start_row = (header_row or 0) + 1
    for row_index, row_values in row_stream:
        if row_index < start_row:
            continue
        values = list(row_values)
        if not any(value not in (None, "") for value in values):
            continue
        data_rows += 1
        padded = values[:total_columns] + [None] * max(0, total_columns - len(values))
        for column_index, value in enumerate(padded):
            kind = value_type(value)
            types[column_index][kind] = types[column_index].get(kind, 0) + 1
            if kind != "empty":
                counts[column_index] += 1
        if len(samples) < 5:
            samples.append({
                "row": row_index,
                "values": [json_value(value) for value in padded],
            })
    return {
        "rows": total_rows,
        "columns": total_columns,
        "headerRow": header_row,
        "headers": [header_name(headers, index) for index in range(total_columns)],
        "dataRows": data_rows,
        "columnProfiles": [
            {
                "index": index + 1,
                "name": header_name(headers, index),
                "nonEmpty": counts[index],
                "types": types[index],
            }
            for index in range(total_columns)
        ],
        "sampleRows": samples,
    }


def inspect_file(root: Path, relative_file: str, sheet_name: str | None) -> dict[str, Any]:
    path = (root / relative_file).resolve(strict=True)
    if root not in path.parents or not path.is_file():
        fail("FILE_OUTSIDE_SOURCE_ROOT")
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        import openpyxl
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            if sheet_name is None:
                fail("SHEET_REQUIRED")
            if sheet_name not in workbook.sheetnames:
                fail("SHEET_NOT_FOUND")
            sheet = workbook[sheet_name]
            header_row, headers = choose_header(sheet.iter_rows(
                min_row=1,
                max_row=min(sheet.max_row, MAX_HEADER_SCAN_ROWS),
                values_only=True,
            ))
            profile = content_profile(
                enumerate(sheet.iter_rows(values_only=True), start=1),
                header_row,
                headers,
                sheet.max_row,
                sheet.max_column,
            )
        finally:
            workbook.close()
    elif suffix == ".xls":
        import xlrd
        workbook = xlrd.open_workbook(path, on_demand=True)
        try:
            if sheet_name is None:
                fail("SHEET_REQUIRED")
            try:
                sheet = workbook.sheet_by_name(sheet_name)
            except xlrd.biffh.XLRDError:
                fail("SHEET_NOT_FOUND")
            header_row, headers = choose_header(
                sheet.row_values(index)
                for index in range(min(sheet.nrows, MAX_HEADER_SCAN_ROWS))
            )
            profile = content_profile(
                ((index + 1, sheet.row_values(index)) for index in range(sheet.nrows)),
                header_row,
                headers,
                sheet.nrows,
                sheet.ncols,
            )
        finally:
            workbook.release_resources()
    elif suffix in {".csv", ".tsv"}:
        delimiter = "," if suffix == ".csv" else "\t"
        rows = None
        encoding_used = None
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                with path.open("r", encoding=encoding, newline="", errors="strict") as handle:
                    rows = list(csv.reader(handle, delimiter=delimiter))
                encoding_used = encoding
                break
            except UnicodeDecodeError:
                continue
        if rows is None:
            fail("UNSUPPORTED_ENCODING")
        header_row, headers = choose_header(rows[:MAX_HEADER_SCAN_ROWS])
        profile = content_profile(
            enumerate(rows, start=1),
            header_row,
            headers,
            len(rows),
            max((len(row) for row in rows), default=0),
        )
        profile["encoding"] = encoding_used
    else:
        fail("UNSUPPORTED_FILE")
    return {
        "rootLabel": root.name,
        "file": relative_file,
        "sheet": sheet_name or "",
        **profile,
    }


def preview_rows(
    rows: Iterable[Iterable[Any]],
    total_rows: int,
    total_columns: int,
) -> dict[str, Any]:
    buffered = [list(row) for row in rows]
    header_row, headers = choose_header(buffered[:MAX_HEADER_SCAN_ROWS])
    start = header_row or 0
    samples = []
    for offset, row in enumerate(buffered[start:start + 3], start=start + 1):
        values = row[:total_columns] + [None] * max(0, total_columns - len(row))
        samples.append({"row": offset, "values": [json_value(value) for value in values]})
    return {
        "rows": total_rows,
        "columns": total_columns,
        "headerRow": header_row,
        "headers": [header_name(headers, index) for index in range(total_columns)],
        "sampleRows": samples,
    }


def tabular_catalog(root: Path) -> dict[str, Any]:
    records = []
    for path in files_under(root):
        suffix = path.suffix.lower()
        if suffix not in SUPPORTED:
            continue
        relative = path.relative_to(root).as_posix()
        record: dict[str, Any] = {"file": relative, "extension": suffix, "sheets": []}
        try:
            if suffix == ".xlsx":
                import openpyxl
                workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
                try:
                    for sheet in workbook.worksheets:
                        rows = list(sheet.iter_rows(
                            min_row=1,
                            max_row=min(sheet.max_row, MAX_HEADER_SCAN_ROWS + 3),
                            values_only=True,
                        ))
                        record["sheets"].append({
                            "name": sheet.title,
                            **preview_rows(rows, sheet.max_row, sheet.max_column),
                        })
                finally:
                    workbook.close()
            elif suffix == ".xls":
                import xlrd
                workbook = xlrd.open_workbook(path, on_demand=True)
                try:
                    for sheet in workbook.sheets():
                        rows = [
                            sheet.row_values(index)
                            for index in range(min(sheet.nrows, MAX_HEADER_SCAN_ROWS + 3))
                        ]
                        record["sheets"].append({
                            "name": sheet.name,
                            **preview_rows(rows, sheet.nrows, sheet.ncols),
                        })
                        workbook.unload_sheet(sheet.name)
                finally:
                    workbook.release_resources()
            else:
                delimiter = "," if suffix == ".csv" else "\t"
                rows = []
                row_count = 0
                encoding_used = None
                for encoding in ("utf-8-sig", "gb18030"):
                    try:
                        with path.open("r", encoding=encoding, newline="", errors="strict") as handle:
                            reader = csv.reader(handle, delimiter=delimiter)
                            for row_count, row in enumerate(reader, start=1):
                                if row_count <= MAX_HEADER_SCAN_ROWS + 3:
                                    rows.append(row)
                        encoding_used = encoding
                        break
                    except UnicodeDecodeError:
                        rows = []
                        row_count = 0
                if encoding_used is None:
                    raise ValueError("UNSUPPORTED_ENCODING")
                columns = max((len(row) for row in rows), default=0)
                record["sheets"].append({
                    "name": "",
                    "encoding": encoding_used,
                    **preview_rows(rows, row_count, columns),
                })
        except Exception as error:
            record["parseError"] = f"{type(error).__name__}: {error}"
        records.append(record)
    return {
        "rootLabel": root.name,
        "fileCount": len(records),
        "sheetCount": sum(len(record["sheets"]) for record in records),
        "files": records,
    }


def inspect_workbook(root: Path, relative_file: str) -> dict[str, Any]:
    path = (root / relative_file).resolve(strict=True)
    if root not in path.parents or not path.is_file():
        fail("FILE_OUTSIDE_SOURCE_ROOT")
    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".xls"}:
        fail("WORKBOOK_REQUIRED")

    sheets = []
    if suffix == ".xlsx":
        import openpyxl
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                rows = list(sheet.iter_rows(
                    min_row=1,
                    max_row=min(sheet.max_row, MAX_HEADER_SCAN_ROWS + 3),
                    values_only=True,
                ))
                sheets.append({
                    "name": sheet.title,
                    **preview_rows(rows, sheet.max_row, sheet.max_column),
                })
        finally:
            workbook.close()
    else:
        import xlrd
        workbook = xlrd.open_workbook(path, on_demand=True)
        try:
            for sheet in workbook.sheets():
                rows = [
                    sheet.row_values(index)
                    for index in range(min(sheet.nrows, MAX_HEADER_SCAN_ROWS + 3))
                ]
                sheets.append({
                    "name": sheet.name,
                    **preview_rows(rows, sheet.nrows, sheet.ncols),
                })
                workbook.unload_sheet(sheet.name)
        finally:
            workbook.release_resources()
    return {
        "rootLabel": root.name,
        "file": relative_file,
        "sheetCount": len(sheets),
        "sheets": sheets,
    }


def search_xlsx(path: Path, term: str, remaining: int) -> list[dict[str, Any]]:
    import openpyxl
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    matches = []
    try:
        for sheet in workbook.worksheets:
            scan_rows = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, MAX_HEADER_SCAN_ROWS), values_only=True))
            header_row, headers = choose_header(scan_rows)
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                for column_index, value in enumerate(row):
                    if value is not None and term in str(value):
                        matches.append({
                            "sheet": sheet.title,
                            "row": row_index,
                            "column": column_index + 1,
                            "field": header_name(headers, column_index),
                            "headerRow": header_row,
                        })
                        if len(matches) >= remaining:
                            return matches
    finally:
        workbook.close()
    return matches


def search_xls(path: Path, term: str, remaining: int) -> list[dict[str, Any]]:
    import xlrd
    workbook = xlrd.open_workbook(path, on_demand=True)
    matches = []
    try:
        for sheet in workbook.sheets():
            header_row, headers = choose_header(
                sheet.row_values(index)
                for index in range(min(sheet.nrows, MAX_HEADER_SCAN_ROWS))
            )
            for row_index in range(sheet.nrows):
                for column_index, value in enumerate(sheet.row_values(row_index)):
                    if term in str(value):
                        matches.append({
                            "sheet": sheet.name,
                            "row": row_index + 1,
                            "column": column_index + 1,
                            "field": header_name(headers, column_index),
                            "headerRow": header_row,
                        })
                        if len(matches) >= remaining:
                            return matches
            workbook.unload_sheet(sheet.name)
    finally:
        workbook.release_resources()
    return matches


def search_delimited(path: Path, term: str, delimiter: str, remaining: int) -> list[dict[str, Any]]:
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="", errors="strict") as handle:
                rows = list(csv.reader(handle, delimiter=delimiter))
            header_row, headers = choose_header(rows[:MAX_HEADER_SCAN_ROWS])
            matches = []
            for row_index, row in enumerate(rows, start=1):
                for column_index, value in enumerate(row):
                    if term in value:
                        matches.append({
                            "sheet": "",
                            "row": row_index,
                            "column": column_index + 1,
                            "field": header_name(headers, column_index),
                            "headerRow": header_row,
                        })
                        if len(matches) >= remaining:
                            return matches
            return matches
        except UnicodeDecodeError:
            continue
    return []


def search(root: Path, term: str, max_matches: int) -> dict[str, Any]:
    matches = []
    errors = []
    all_files = files_under(root)
    tabular_files = [path for path in all_files if path.suffix.lower() in SUPPORTED]
    per_file_limit = max(1, max_matches // max(1, len(tabular_files)))
    truncated_files = []
    for file_index, path in enumerate(all_files, start=1):
        suffix = path.suffix.lower()
        if suffix not in SUPPORTED:
            continue
        try:
            if suffix == ".xlsx":
                located = search_xlsx(path, term, per_file_limit)
            elif suffix == ".xls":
                located = search_xls(path, term, per_file_limit)
            else:
                located = search_delimited(
                    path,
                    term,
                    "," if suffix == ".csv" else "\t",
                    per_file_limit,
                )
            relative = path.relative_to(root).as_posix()
            if len(located) >= per_file_limit:
                truncated_files.append(f"F{file_index:04d}")
            matches.extend(
                {
                    "fileId": f"F{file_index:04d}",
                    "file": safe_relative_path(relative),
                    **item,
                }
                for item in located
            )
        except Exception as error:
            errors.append({
                "fileId": f"F{file_index:04d}",
                "file": safe_relative_path(path.relative_to(root).as_posix()),
                "error": type(error).__name__,
            })
    return {
        "rootLabel": root.name,
        "matchCount": len(matches),
        "sampledPerFile": per_file_limit,
        "truncated": bool(truncated_files),
        "truncatedFileIds": truncated_files,
        "matches": matches,
        "errors": errors,
    }


def write_json(output: Path, payload: dict[str, Any]) -> None:
    temporary = output.with_suffix(output.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")
    os.replace(temporary, output)
    print(json.dumps({
        "status": "completed",
        "output": str(output),
        "fileCount": payload.get("fileCount"),
        "matchCount": payload.get("matchCount"),
        "truncated": payload.get("truncated", False),
    }, ensure_ascii=False))


def main() -> None:
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)
    inventory_parser = subparsers.add_parser("inventory")
    inventory_parser.add_argument("--root", required=True)
    inventory_parser.add_argument("--output", required=True)
    search_parser = subparsers.add_parser("search")
    search_parser.add_argument("--root", required=True)
    search_parser.add_argument("--term", required=True)
    search_parser.add_argument("--output", required=True)
    search_parser.add_argument("--max-matches", type=int, default=500)
    inspect_parser = subparsers.add_parser("inspect")
    inspect_parser.add_argument("--root", required=True)
    inspect_parser.add_argument("--file", required=True)
    inspect_parser.add_argument("--sheet")
    inspect_parser.add_argument("--output", required=True)
    workbook_parser = subparsers.add_parser("workbook")
    workbook_parser.add_argument("--root", required=True)
    workbook_parser.add_argument("--file", required=True)
    workbook_parser.add_argument("--output", required=True)
    catalog_parser = subparsers.add_parser("catalog")
    catalog_parser.add_argument("--root", required=True)
    catalog_parser.add_argument("--output", required=True)
    args = parser.parse_args()
    root, output = safe_paths(args.root, args.output)
    if args.command == "inventory":
        payload = inventory(root)
    elif args.command == "search":
        if not args.term.strip():
            fail("SEARCH_TERM_REQUIRED")
        if args.max_matches < 1 or args.max_matches > 10_000:
            fail("MAX_MATCHES_OUT_OF_RANGE")
        payload = search(root, args.term, args.max_matches)
    elif args.command == "inspect":
        payload = inspect_file(root, args.file, args.sheet)
    elif args.command == "workbook":
        payload = inspect_workbook(root, args.file)
    else:
        payload = tabular_catalog(root)
    write_json(output, payload)


if __name__ == "__main__":
    main()
